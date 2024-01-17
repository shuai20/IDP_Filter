import openpyxl

# Load the workbook and select the active worksheet
workbook = openpyxl.load_workbook('/mnt/data/Finalall.xlsx')
sheet = workbook.active

# Define the keywords for the two groups
keywords_group_1 = [
    "Profile & Contact Information",
    "Product Usage",
    "Device Information",
    "Participant Profile & Contact Information",
    "Participants"
]

keywords_group_2 = [
    "Content",
    "Calendars",
    "Registration & Scheduling",
    "Devices"
]

# Function to search keywords in a row for a specific group
def search_keywords(row, keywords):
    for cell in row:
        if cell.value:  # Check if the cell has a value to prevent errors
            for keyword in keywords:
                if keyword.lower() in cell.value.lower():
                    return 1
    return 0

# Apply the function to each row and write the result in the Q and R columns
for row in sheet.iter_rows(min_row=2, max_row=1842, min_col=2, max_col=16):
    result_group_1 = search_keywords(row, keywords_group_1)
    result_group_2 = search_keywords(row, keywords_group_2)
    sheet.cell(row=row[0].row, column=17, value=result_group_1)  # Q column for group 1
    sheet.cell(row=row[0].row, column=18, value=result_group_2)  # R column for group 2

# Save the modified workbook to a new file
modified_file_path = '/mnt/data/modified_Finalall.xlsx'
workbook.save(modified_file_path)


